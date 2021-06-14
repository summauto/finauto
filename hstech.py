import yfinance as yf
import os
import time

cwd = os.getcwd()
SHEET = "HSTECH" #Sheet
TICKER_COLUMN = 1
PRICE_COLUMN = 2

def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()

def getTickers(cwd):
    from openpyxl import load_workbook
    # for first column, and how many rows where there is data
        #take these as ticker names
    workbook = load_workbook(filename =cwd+"/stockPrices.xlsx")
    sheet = workbook[SHEET]
    i = 2
    myTickers = []
    while sheet.cell(row=i, column=TICKER_COLUMN).value != None:
        myTickers.append(sheet.cell(row=i, column=TICKER_COLUMN).value)
        i+=1
    else: 
        print("i have collected "+str(len(myTickers))+" tickers.")

    return myTickers
    

def priceList(myTickers):
    #put the ticker names here and get the last closing price
    myInfo = [[],[]]
    # myInfo = []
    print("getting prices...")
    l = len(myTickers)
    printProgressBar(0, l, prefix = 'Progress:', suffix = 'Complete', length = 50)

    """
    info dataframe:
    {'zip', 'sector', 'fullTimeEmployees', 'longBusinessSummary', 'city', 'phone', 'state', 'country', 'companyOfficers', 'website', 'maxAge',
    'address1', 'industry', 'previousClose', 'regularMarketOpen', 'twoHundredDayAverage', 'trailingAnnualDividendYield', 'payoutRatio',
    'volume24Hr', 'regularMarketDayHigh', 'navPrice', 'averageDailyVolume10Day', 'totalAssets', 'regularMarketPreviousClose', 'fiftyDayAverage',
    'trailingAnnualDividendRate', 'open', 'toCurrency', 'averageVolume10days', 'expireDate', 'yield', 'algorithm', 'dividendRate', 'exDividendDate'
    'beta', 'circulatingSupply', 'startDate', 'regularMarketDayLow', 'priceHint', 'currency', 'trailingPE', 'regularMarketVolume', 'lastMarket',
    'maxSupply', 'openInterest', 'marketCap', 'volumeAllCurrencies', 'strikePrice', 'averageVolume', 'priceToSalesTrailing12Months', 'dayLow',
    'ask', 'ytdReturn', 'askSize', 'volume', 'fiftyTwoWeekHigh', 'forwardPE', 'fromCurrency', 'fiveYearAvgDividendYield', 'fiftyTwoWeekLow', 
    'bid', 'tradeable', 'dividendYield', 'bidSize', 'dayHigh', 'exchange', 'shortName', 'longName', 'exchangeTimezoneName',
    'exchangeTimezoneShortNaalHoldingsTurnover', 'enterpriseToRevenue', 'beta3Year', 'profitMargins', 'enterpriseToEbitda', '52WeekChange', 
    'morningStarRiskRating', 'forwardEps', 'revenueQuarterlyGrowth' (unreliable), 'sharesOutstanding', 'fundInceptionDate' (none),
    'annualReportExpenseRatio' (none), 'bookValue', 'sharesShort', 'sharesPercentSharesOut', 'fundFamily', 'lastFiscalYearEnd',
    'heldPercentInstitutions', 'netIncomeToCommon', 'trailingEps', 'lastDividendValue', 'SandP52WeekChange', 'priceToBook', 'heldPercentInsiders',
    'nextFiscalYearEnd',  'mostRecentQuarter','shortRatio','sharesShortPreviousMonthDate','floatShares', 'enterpriseValue', 'threeYearAverageReturn',
    'lastSplitDate', 'lastSplitFactor','legalType', 'lastDividendDate', 'morningStarOverallRating', 'earningsQuarterlyGrowth', 'dateShortInterest',
    'pegRatio', 'lastCapGain', 'shortPercentOfFloat', 'sharesShortPriorMonth', 'impliedSharesOutstanding', 'category', 'fiveYearAverageReturn',
    'regularMarketPrice', 'logo_url'}
    """

    for i in range(len(myTickers)):
        tickerList = yf.Ticker(str(myTickers[i]))
        #print("getting for ticker: "+str(myTickers[i])+" = ")
        printProgressBar(i + 1, len(myTickers), prefix = 'Progress:', suffix = 'Complete', length = 50)
        try: 
            previousClose = tickerList.info["previousClose"]
            myInfo[0].append(previousClose)
            # print(str(previousClose)+"\n")
        except:
            myInfo[0].append("NIL")
            # print("NIL\n")
        # try:
        #     floatShares = tickerList.info["floatShares"]
        #     myInfo[1].append(floatShares)
        # except:
        #     myInfo[1].append("NIL")
        #     # print("NIL\n")
        
    return myInfo

    ### note: im getting an issue pulling previousClose for 3067.hk??

    ###### this works in getting previous closing price of a hk stock  #############
    # tencent = yf.Ticker("BTC-US")
    # print(tencent.info)#["previousClose"])
    

def printExcel(price):
    from openpyxl import Workbook, load_workbook 
    from openpyxl.comments import Comment
    workbook = load_workbook(filename =cwd+"/stockPrices.xlsx")
    sheet = workbook[SHEET]
    print("inserting prices...")

    for i in range(len(price)):
        for j in range(len(price[i])):
            #sheet.cell(row=i+1, column=3).value = i
            if price[i][j] != "NIL":
                sheet.cell(row=j+2, column=PRICE_COLUMN).value = price[i][j]
            else: 
                sheet.cell(row=j+2, column=PRICE_COLUMN).comment = Comment("This cell was not updated","finsbot")
            #print("ive added "+str(price[i])+" to "+str(sheet.cell(row=i+2, column=3)))
    workbook.save(filename=cwd+"/stockPrices.xlsx")

if __name__ == '__main__':
    printExcel(priceList(getTickers(cwd)))